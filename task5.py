from networkx import *
import pylab as plt
from math import sqrt
from matplotlib import pyplot
from openpyxl import load_workbook


# ====== Открываем книгу Excel и считываем из нее данные. Считываем до тех пор, пока не встретится
# ======# в книге пустая ячейка. Данные представляют собой ребра графа и их вес. Данные сохраняем в файл Edges
def GetEdges():
    wb = load_workbook('./data.xlsx')
    sheet = wb.active
    file = open('Edges.txt', 'wt')
    i = 2
    check1 = True
    while (check1):
        if (sheet.cell(row=i, column=1).value) is None:
            check1 = False
        else:
            file.write(str(sheet.cell(row=i, column=1).value) + '\t' +
                       str(sheet.cell(row=i, column=2).value) + '\t' +
                       str(sheet.cell(row=i, column=7).value) + '\n')
            i = i + 1
    file.close()


# ===== Аналогично открываем книгу Excel и считываем данные. Считываем все вершины из 1 колонки и их координаты,
# ===== затем из второй. Сохраняем данные в файл и в словарь pos. На всякий случай создаем список labels с именами вершин
def GetVertex():
    file = open('Vertex.txt', 'wt')
    global labels
    global pos
    wb = load_workbook('./data.xlsx')
    sheet = wb.active
    pos = {}
    i = 2
    check1 = True
    while (check1):
        if (sheet.cell(row=i, column=1).value) is None:
            check1 = False
        else:
            vert = int(sheet.cell(row=i, column=1).value)
            pos[vert] = (float(sheet.cell(row=i, column=3).value),
                         float(sheet.cell(row=i, column=4).value))
            file.write(str(sheet.cell(row=i, column=1).value) + '\t' +
                       str(sheet.cell(row=i, column=3).value) + '\t' +
                       str(sheet.cell(row=i, column=4).value) + '\n')
            i = i + 1
    i = 2
    check1 = True
    while (check1):
        if (sheet.cell(row=i, column=1).value) is None:
            check1 = False
        else:
            vert = int(sheet.cell(row=i, column=2).value)
            pos[vert] = (float(sheet.cell(row=i, column=5).value),
                         float(sheet.cell(row=i, column=6).value))
            file.write(str(sheet.cell(row=i, column=2).value) + '\t' +
                       str(sheet.cell(row=i, column=5).value) + '\t' +
                       str(sheet.cell(row=i, column=6).value) + '\n')
            i = i + 1
    labels = {}
    for i in pos.keys():
        labels[i] = i
    file.close()


# ====Обработка файла с координатами входных двух точек и проверка границ
def CheckInput():
    file = open('coordinates.txt', 'r')
    s = file.read()
    file.close()
    ss = s.split('\t')
    a = (float(ss[1]), float(ss[2].split('\n')[0]))
    b = (float(ss[3]), float(ss[4].split('\n')[0]))
    check1 = False
    if (a[0] <= 38.9774837690821) or (a[0] >= 39.0283642309179):
        print('Нарушены границы допустимого значения X в точке А')
        check1 = True
    if (a[1] <= 45.0549145480683) or (a[1] >= 45.0908534519317):
        print('Нарушены границы допустимого значения Y в точке А')
        check1 = True
    if (b[0] <= 38.9774837690821) or (a[0] >= 39.0283642309179):
        print('Нарушены границы допустимого значения X в точке B')
        check1 = True
    if (b[1] <= 45.0549145480683) or (a[1] >= 45.0908534519317):
        print('Нарушены границы допустимого значения Y в точке B')
        check1 = True
    if (check1):
        print('Измените координаты!')
    else:
        print('Точки добавлены!')
    return a, b


# ======Отображает основные характеристики графа
def GetGraphInfo(gr):
    s = ''
    ss = nx.info(gr).split('\n')
    ss[0] = 'Анализ графа\n'
    ss[1] = 'Тип:\t\tГраф'
    ss[2] = 'Количество вершин: \t ' + ss[2][17:len(ss[2])]
    ss[3] = 'Количество ребер: \t ' + ss[3][17:len(ss[3])]
    for i in range(ss.__len__() - 1):
        s = s + ss[i] + '\n'
    return s


# ======Выводит ребра и их вес
def ShowEdgesWeight(gr):
    return gr.edges(data=True)


# ======Находит дистанцию между двумя точками. Нужно, чтобы найти ближайшие вершины к заданным точкам
def Dist(a, b):
    return sqrt((b[0] - a[0]) ** 2 + (b[1] - a[1]) ** 2)


# =====Визуализация графа и нахождение кратчайшего пути
def PaintGraph(gr, a, b):  # Сначала находим ближайшие вершины к заданным точкам А и В
    min_dist_near_a = Dist(a, pos.get(list(gr.nodes())[10]))  # Обозначаем для поиска наиближайшей точки любую,
    min_dist_near_b = Dist(b, pos.get(list(gr.nodes())[10]))  # для того чтобы найти минимум дистанции
    dot_near_a = pos.get(list(gr.nodes())[10])  # Обозначаем минимальной дистанцией любую дистанцию между точками
    dot_near_b = pos.get(list(gr.nodes())[10])  # В данном случае взята 11 по счету вершина ближайшей
    for i in pos:
        dot = pos.get(i)
        d = Dist(dot, a)  # Находим ближайшую вершину к точке А и расстояние и между ними, которое
        if (d < min_dist_near_a):  # является минимальным
            min_dist_near_a = d
            dot_near_a = i
        d = Dist(dot, b)
        if (d < min_dist_near_b):  # Находим ближайшую вершину к точке В и расстояние и между ними, которое
            min_dist_near_b = d  # является минимальным
            dot_near_b = i
    gr.add_node(1)  # Добавляем введенные точки А и В к графу, чтобы потом визуализировать их
    gr.add_node(2)
    pos[1] = a  # Добавляем их координаты
    pos[2] = b

    # Для поиска кратчайшего пути используем алгоритм Дейкстры
    pathfind = nx.dijkstra_path(gr, dot_near_a, dot_near_b, 'weight')  # Метод выдает список, состоящей из вершин,
    # которые входят в кратчайший путь
    path_len = nx.dijkstra_path_length(gr, dot_near_a, dot_near_b, 'weight')  # Выводит длину пути. Высчитывает тем же
    # алгоритмом Дейкстры
    # Получаем ребра графа, узлами которых являются вершины пути с помощью метода edges и переводим в тип списка
    # Находим узлы, которые граничат с узлами пути и удаляем связывающие их ребра. Таким образом, удаляем лишние ребра из списка
    print('Кратчайший путь проходит через ', pathfind.__len__(), ' вершин')
    print('Длина кратчайшего пути, найденная методом Дейкстры: ', path_len)
    path_edges = []
    for i in list(gr.edges(pathfind, data=False)):
        if (i[0] in pathfind) and (i[1] in pathfind):
            path_edges.append(i)
            # Отрисовка графа
    nx.draw_networkx(gr, pos, with_labels=False, node_color='b', node_size=5, alpha=0.8,
                     width=2, font_size=7)

    # Закрашивание вершин, участвующих в пути красным
    nx.draw_networkx_nodes(gr, pos, nodelist=pathfind, node_color='r', node_size=20, alpha=0.8)
    # Закрашивание введенных точек желтым
    nx.draw_networkx_nodes(gr, pos, nodelist=[1, 2], node_color='y', node_size=30, alpha=0.8)
    # Закрашивание кратчайшего пути красным
    nx.draw_networkx_edges(gr, pos, width=2.5, alpha=0.8, edgelist=path_edges, edge_color='r')
    # Визуализация графа с помощью matplotlib
    plt.axis('on')
    plt.title('Map')
    plt.show()
    plt.savefig('Map.png')


# =========================================================================================================
GetEdges()  # Записываем данные о ребрах в txt файл
G = nx.read_weighted_edgelist("Edges.txt", nodetype=int)  # Строим граф, считывая данные из файла
GetVertex()  # Получаем координаты вершин графа
print(GetGraphInfo(G))  # Выводим информацию о графе
print('В дирректории программы находится файл coordinates.txt.\n'
      'Чтобы ввести свои значения точек А и В измените значения, данные в файле.')
A, B = CheckInput()
PaintGraph(G, A, B)
